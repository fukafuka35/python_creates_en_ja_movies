import openpyxl
from moviepy.editor import *
from gtts import gTTS
import os

# メモリの解放を明示的に行う　20230827　ガベージコレクション
import gc

# 20230826 深澤修正 (再帰？)関数の実行回数の上限を変更
import sys
import threading
sys.setrecursionlimit(67108864)
threading.stack_size(1024*1024)

# 深澤修正 カレントディレクトリを変更   (rをつけるとバックスラッシュを使える)
#path = os.getcwd()
#print(path)
os.chdir(r'C:\Users\fukas\Desktop\くーくん 英語 聞き流し')
#path2 = os.getcwd()
#print(path2)
#exit()

# Excelファイルの読み込み
wb = openpyxl.load_workbook('英語リスト.xlsx')
sheet = wb.active

# 動画のフレームレートとサイズの設定
fps = 30
width = 1080
height = 1920

# 動画ファイルの保存パスと一時フォルダの設定
output_path = 'output.mp4'
temp_folder = 'temp'

# 一時フォルダの作成
os.makedirs(temp_folder, exist_ok=True)


# オープニングのスライド
slide_open2 = ImageClip('背景2.png').set_duration('00:00:05.00')

txt_clip_op1 = TextClip('英単語 聞き流し', fontsize=120, color='white', font='C:/Windows/Fonts/YuGothB.ttc')
txt_clip_op1 = txt_clip_op1.set_position((110, 290)).set_start(1).set_duration('00:00:3.00')
slide_open2 = CompositeVideoClip([slide_open2, txt_clip_op1])
txt_clip_op2 = TextClip('名詞', fontsize=100, color='white', font='C:/Windows/Fonts/YuGothB.ttc')
txt_clip_op2 = txt_clip_op2.set_position((110, 600)).set_start(2).set_duration('00:00:02.00')
slide_open2 = CompositeVideoClip([slide_open2, txt_clip_op2])

final_video = slide_open2
# final_video = concatenate_videoclips([slide_open2, slide_open3])


# 各スライドごとに処理を行う
for i in range(2, 62):

    # 英語のテキストを取得
    text_en = sheet.cell(row=i, column=1).value

    # 日本語のテキストを取得
    text_ja = sheet.cell(row=i, column=2).value

    # スライド準備 最初
    slide_first = ImageClip('背景2.png').set_duration('00:00:10.00')


    # スライドにテキストを重ねる 最初
    telop_en = TextClip(text_en, fontsize=130, color='white', font='Arial', align='West')
    telop_en = telop_en.set_position((120, 290)).set_duration('00:00:09.00')
    slide_first = CompositeVideoClip([slide_first, telop_en])
    telop_ja = TextClip(text_ja, fontsize=90, color='white', font='C:/Windows/Fonts/YuGothB.ttc')
    telop_ja = telop_ja.set_position((120, 700)).set_start(1).set_duration('00:00:08.00')
    slide_first = CompositeVideoClip([slide_first, telop_ja])
#    explain_text_en2 = explain_text_en.replace('  ', '\n')


    # 英語の音声を作る
    voice_en = gTTS(text_en, lang='en')
    voice_en.save(os.path.join(temp_folder, 'voice_{:03d}_en.mp3'.format(i)))

   # 20231002　音を作る 
    voice_ja = gTTS(text_ja, lang='ja')
    voice_ja.save(os.path.join(temp_folder, 'voice_{:03d}_ja.mp3'.format(i)))

    # スライドに音を重ねる
    audio_en = AudioFileClip(os.path.join(temp_folder, 'voice_{:03d}_en.mp3'.format(i)))
    audio_ja = AudioFileClip(os.path.join(temp_folder, 'voice_{:03d}_ja.mp3'.format(i)))
    audio_first = concatenate_audioclips([audio_en, audio_ja, audio_en, audio_ja])
    slide_first = slide_first.set_audio(audio_first)


    final_video = concatenate_videoclips([final_video, slide_first])

    # メモリの解放を明示的に行う　20230827　ガベージコレクション　ここに書いても4行エクセルは正常に処理できた
    gc.collect


# 締めのご挨拶 3
slide_end3 = ImageClip('背景2.png').set_duration('00:00:10.00')

txt_clip_end3 = TextClip('ご視聴\nありがとう\nございました', fontsize=120, color='white', font='C:/Windows/Fonts/YuGothB.ttc')
txt_clip_end3 = txt_clip_end3.set_position(('center', 580)).set_duration('00:00:10.00')
slide_end3 = CompositeVideoClip([slide_end3, txt_clip_end3])


final_video = concatenate_videoclips([final_video, slide_end3])

# 動画ファイルの保存
final_video.write_videofile(output_path, fps=fps, codec='libx264')
